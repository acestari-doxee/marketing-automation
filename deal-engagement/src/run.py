"""
Deal Engagement Extractor — HubSpot → Excel
Reproduces the format of hubspot_deal_report.xlsx.

Usage:
    python run.py            # normal run
    python run.py --init     # prints the mapping for the cached JSON files
    python run.py --no-cache # ignore the cache, redo all API calls

Output: output/hubspot_deal_report_<YYYYMMDD_HHmm>.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Costanti / setup
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "config.yaml"
CACHE_DIR = ROOT / "data" / "cache"
OUTPUT_DIR = ROOT / "output"

HUBSPOT_BASE = "https://api.hubapi.com"

CACHE_FILES = {
    1: "step1_list_memberships.json",
    2: "step2_deals_detail.json",
    3: "step3_deal_to_companies.json",
    4: "step4_companies_detail.json",
    5: "step5_company_to_contacts.json",
    6: "step6_contacts_detail.json",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_config() -> dict:
    if not CONFIG_PATH.exists():
        sys.exit(f"[ERR] config.yaml not found in {CONFIG_PATH}")
    with CONFIG_PATH.open() as f:
        cfg = yaml.safe_load(f)
    cfg.setdefault("score_property", "lead_score_contacts_total")
    cfg.setdefault("score_threshold", None)
    cfg.setdefault("output_dir", str(OUTPUT_DIR))
    cfg.setdefault("cache_dir", str(CACHE_DIR))
    return cfg


def get_token() -> str:
    load_dotenv(ROOT / ".env")
    tok = os.getenv("HUBSPOT_TOKEN")
    if not tok:
        sys.exit("[ERR] HUBSPOT_TOKEN missing. Copy .env.example to .env and paste the token.")
    return tok


def hs_request(
    method: str,
    path: str,
    token: str,
    json_body: dict | None = None,
    retries: int = 3,
) -> dict:
    """HubSpot request with retry on 429/5xx."""
    url = f"{HUBSPOT_BASE}{path}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    for attempt in range(retries):
        r = requests.request(method, url, headers=headers, json=json_body, timeout=30)
        if r.status_code == 429 or r.status_code >= 500:
            wait = 2 ** attempt
            print(f"  [warn] {r.status_code} — retry in {wait}s")
            time.sleep(wait)
            continue
        r.raise_for_status()
        return r.json()
    sys.exit(f"[ERR] {method} {path} failed after {retries} attempts")


def cache_path(step: int) -> Path:
    return Path(CACHE_DIR) / CACHE_FILES[step]


def load_or_fetch(step: int, fetcher, *, force: bool) -> dict:
    """If the cache file exists and force=False, read it. Otherwise call fetcher() and write it."""
    path = cache_path(step)
    if not force and path.exists():
        print(f"  [cache] step {step} ← {path.name}")
        with path.open() as f:
            return json.load(f)
    print(f"  [api]   step {step} → HubSpot")
    data = fetcher()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w") as f:
        json.dump(data, f, indent=2)
    return data


def chunked(lst: list, n: int):
    for i in range(0, len(lst), n):
        yield lst[i : i + n]


# ---------------------------------------------------------------------------
# Step API
# ---------------------------------------------------------------------------
def step1_list_memberships(token: str, list_id: int) -> dict:
    return hs_request(
        "GET", f"/crm/v3/lists/{list_id}/memberships/join-order", token
    )


def step2_deals_detail(token: str, deal_ids: list[str]) -> dict:
    body = {
        "inputs": [{"id": did} for did in deal_ids],
        "properties": ["dealname", "amount", "salesforce_create_date", "dealstage", "pipeline"],
    }
    return hs_request("POST", "/crm/v3/objects/deals/batch/read", token, body)


def step3_deal_to_companies(token: str, deal_ids: list[str]) -> dict:
    body = {"inputs": [{"id": did} for did in deal_ids]}
    return hs_request(
        "POST", "/crm/v4/associations/deals/companies/batch/read", token, body
    )


def step4_companies_detail(token: str, company_ids: list[str]) -> dict:
    body = {
        "inputs": [{"id": cid} for cid in company_ids],
        "properties": ["name", "domain"],
    }
    return hs_request("POST", "/crm/v3/objects/companies/batch/read", token, body)


def step5_company_to_contacts(token: str, company_ids: list[str]) -> dict:
    body = {"inputs": [{"id": cid} for cid in company_ids]}
    return hs_request(
        "POST", "/crm/v4/associations/companies/contacts/batch/read", token, body
    )


def step6_contacts_detail(
    token: str, contact_ids: list[str], score_property: str
) -> dict:
    """HubSpot batch read accepts max 100 inputs per call. We chunk them."""
    merged = {"results": []}
    for chunk in chunked(contact_ids, 100):
        body = {
            "inputs": [{"id": cid} for cid in chunk],
            "properties": [
                "firstname",
                "lastname",
                "jobtitle",
                "email",
                score_property,
                "hs_analytics_source",
                "hs_analytics_source_data_1",
                "hs_analytics_source_data_2",
                "first_touch_campaign",
            ],
        }
        out = hs_request("POST", "/crm/v3/objects/contacts/batch/read", token, body)
        merged["results"].extend(out.get("results", []))
    return merged


# ---------------------------------------------------------------------------
# Aggregazione
# ---------------------------------------------------------------------------
def aggregate(
    deals: dict,
    deal_to_company: dict,
    companies: dict,
    company_to_contacts: dict,
    contacts: dict,
    score_property: str,
    score_threshold,
) -> list[dict]:
    """
    Returns a list of groups:
        [{deal: {...}, account_name: str, contacts: [ {name, jobtitle, email, score}, ...]}, ...]
    Sorted by ascending close_date (nearest deals on top).
    """
    # id → record maps
    deal_map = {d["id"]: d["properties"] for d in deals["results"]}
    company_map = {c["id"]: c["properties"] for c in companies["results"]}
    contact_map = {c["id"]: c["properties"] for c in contacts["results"]}

    # deal → company id
    d2c = {}
    for r in deal_to_company["results"]:
        if r["to"]:
            d2c[r["from"]["id"]] = str(r["to"][0]["toObjectId"])

    # company → list of contact ids
    c2contacts = {}
    for r in company_to_contacts["results"]:
        cid = r["from"]["id"]
        c2contacts[cid] = [str(t["toObjectId"]) for t in r["to"]]

    groups = []
    for did, dprops in deal_map.items():
        comp_id = d2c.get(did)
        comp = company_map.get(comp_id, {}) if comp_id else {}
        account_name = comp.get("name") or _account_name_from_dealname(dprops.get("dealname", ""))

        contact_ids = c2contacts.get(comp_id, []) if comp_id else []
        contact_rows = []
        for cid in contact_ids:
            cp = contact_map.get(cid)
            if not cp:
                continue
            score_raw = cp.get(score_property)
            score_num = _to_int(score_raw)

            # filter: always exclude unscored contacts (None/"")
            if score_num is None:
                continue
            # se threshold definito, applica
            if score_threshold is not None and score_num < score_threshold:
                continue

            contact_rows.append(
                {
                    "name": _full_name(cp),
                    "jobtitle": cp.get("jobtitle") or "",
                    "email": cp.get("email") or "",
                    "score": score_num,
                    "source": cp.get("hs_analytics_source") or "",
                    "source_data_1": cp.get("hs_analytics_source_data_1") or "",
                    "source_data_2": cp.get("hs_analytics_source_data_2") or "",
                    "first_touch_campaign": cp.get("first_touch_campaign") or "",
                }
            )

        # sort contacts by score desc
        contact_rows.sort(key=lambda x: x["score"], reverse=True)

        groups.append(
            {
                "deal_id": did,
                "deal_name": dprops.get("dealname", ""),
                "amount": _to_float(dprops.get("amount")),
                "close_date": _fmt_date(dprops.get("salesforce_create_date")),
                "close_date_raw": dprops.get("salesforce_create_date"),
                "account_name": account_name,
                "contacts": contact_rows,
            }
        )

    # sort deals by close_date
    groups.sort(key=lambda g: g["close_date_raw"] or "9999")
    return groups


def _full_name(props: dict) -> str:
    fn = (props.get("firstname") or "").strip()
    ln = (props.get("lastname") or "").strip()
    return f"{fn} {ln}".strip() or (props.get("email") or "")


def _to_int(v):
    if v in (None, "", "—"):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _to_float(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _fmt_date(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except ValueError:
        return iso[:10]


def _account_name_from_dealname(dealname: str) -> str:
    """Fallback when the company has no name: take the first part of the deal name."""
    for sep in [" - ", "_"]:
        if sep in dealname:
            return dealname.split(sep)[0].strip()
    return dealname or "(unknown)"


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
HEADER = [
    "Account Name",
    "Opportunity Name",
    "Amount (€)",
    "SF Create Date",
    "Contact Name",
    "Job Title",
    "Email",
    "Lead Score",
    "Analytics Source",
    "Source Data 1",
    "Source Data 2",
    "First Touch Campaign",
]


def write_xlsx(groups: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Report"

    # header
    for col, h in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for g in groups:
        contacts = g["contacts"]
        if not contacts:
            # deal with no engaged contacts: one row with empty contact fields
            _write_row(
                ws, row,
                g["account_name"], g["deal_name"], g["amount"], g["close_date"],
                "", "", "", "—", "", "", "", "",
            )
            row += 1
            continue

        first = True
        for c in contacts:
            _write_row(
                ws, row,
                g["account_name"] if first else "",
                g["deal_name"] if first else "",
                g["amount"] if first else "",
                g["close_date"] if first else "",
                c["name"], c["jobtitle"], c["email"], c["score"],
                c["source"], c["source_data_1"], c["source_data_2"],
                c["first_touch_campaign"],
            )
            row += 1
            first = False

    # column widths
    widths = [28, 38, 12, 12, 24, 32, 32, 12, 24, 24, 24, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Foglio 2 — Legend
    legend = wb.create_sheet("Legend")
    legend["A1"] = "Lead Score Reference"
    legend["A1"].font = Font(bold=True, size=14)
    rows = [
        ("≥ 10", "High engagement"),
        ("5–9", "Medium engagement"),
        ("0–4", "Low engagement"),
        ("—", "Not yet scored"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        legend.cell(row=i, column=1, value=k).font = Font(bold=True)
        legend.cell(row=i, column=2, value=v)
    legend.column_dimensions["A"].width = 12
    legend.column_dimensions["B"].width = 32

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_row(ws, row, account, opp, amount, close_date, name, job, email, score,
               source, source_data_1, source_data_2, first_touch_campaign):
    ws.cell(row=row, column=1, value=account)
    ws.cell(row=row, column=2, value=opp)
    amount_cell = ws.cell(row=row, column=3, value=amount if amount is not None else "")
    if isinstance(amount, (int, float)):
        amount_cell.number_format = "#,##0"
    ws.cell(row=row, column=4, value=close_date)
    ws.cell(row=row, column=5, value=name)
    ws.cell(row=row, column=6, value=job)
    ws.cell(row=row, column=7, value=email)
    ws.cell(row=row, column=8, value=score)
    ws.cell(row=row, column=9, value=source)
    ws.cell(row=row, column=10, value=source_data_1)
    ws.cell(row=row, column=11, value=source_data_2)
    ws.cell(row=row, column=12, value=first_touch_campaign)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-cache", action="store_true", help="Ignore the cache, redo all calls.")
    ap.add_argument("--init", action="store_true", help="Print the cache file mapping and exit.")
    args = ap.parse_args()

    if args.init:
        print("To reuse your 6 JSON files, copy them into data/cache/ with these names:")
        for k, v in CACHE_FILES.items():
            print(f"  step {k}: {v}")
        return

    cfg = load_config()
    token = get_token()
    force = args.no_cache

    print(f"== Deal Engagement Extractor — list {cfg['list_id']} ==")
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    # 1. memberships
    mem = load_or_fetch(1, lambda: step1_list_memberships(token, cfg["list_id"]), force=force)
    deal_ids = [str(r["recordId"]) for r in mem["results"]]
    print(f"  → {len(deal_ids)} deals in the segment")

    # 2. deal detail
    deals = load_or_fetch(2, lambda: step2_deals_detail(token, deal_ids), force=force)

    # 3. deal → company
    d2c = load_or_fetch(3, lambda: step3_deal_to_companies(token, deal_ids), force=force)
    company_ids = list({str(t["toObjectId"]) for r in d2c["results"] for t in r["to"]})
    print(f"  → {len(company_ids)} unique companies")

    # 4. company detail
    companies = load_or_fetch(
        4, lambda: step4_companies_detail(token, company_ids), force=force
    )

    # 5. company → contacts
    c2c = load_or_fetch(
        5, lambda: step5_company_to_contacts(token, company_ids), force=force
    )
    contact_ids = list({str(t["toObjectId"]) for r in c2c["results"] for t in r["to"]})
    print(f"  → {len(contact_ids)} total contacts across companies")

    # 6. contacts detail (with score)
    contacts = load_or_fetch(
        6,
        lambda: step6_contacts_detail(token, contact_ids, cfg["score_property"]),
        force=force,
    )

    # 7. aggregate + filter
    groups = aggregate(
        deals,
        d2c,
        companies,
        c2c,
        contacts,
        cfg["score_property"],
        cfg["score_threshold"],
    )

    total_engaged = sum(len(g["contacts"]) for g in groups)
    print(f"  → {total_engaged} contacts with a valid score (threshold={cfg['score_threshold']})")

    # 8. xlsx
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = Path(cfg["output_dir"]) / f"hubspot_deal_report_{ts}.xlsx"
    write_xlsx(groups, out_path)
    print(f"\n[OK] Output: {out_path}")


if __name__ == "__main__":
    main()
