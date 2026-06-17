#!/usr/bin/env python3
"""
Event Mailer — RSVP control panel via Microsoft Graph.

Reads live statuses directly from Exchange. Works independently of the local
Outlook client. Compatible with Gmail as a recipient: invites in RFC 5546
format, replies read via the API.

Start with: ./start.command
"""

import json
import threading
import time
import webbrowser
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from zoneinfo import ZoneInfo

import graph_client as gc
import setup

_LOCAL_TZ = ZoneInfo("Europe/Rome")


BASE_DIR    = Path(__file__).parent
INVITATI    = BASE_DIR / "invitati.json"
CONFIG_FILE = BASE_DIR / "config.json"
PORT        = 8765

STATUS_COLORS = {
    "Accepted":  "#22c55e",
    "Declined":  "#ef4444",
    "Tentative": "#eab308",
    "Pending":   "#9ca3af",
}


def _load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def _event_kw():
    return _load_config().get("event_keyword", "DOXEE")


def _fmt_iso(iso_str):
    if not iso_str:
        return ""
    try:
        if "." in iso_str:
            iso_str = iso_str.split(".")[0]
        dt = datetime.fromisoformat(iso_str.replace("Z", "")).replace(tzinfo=timezone.utc)
        return dt.astimezone(_LOCAL_TZ).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return iso_str


def read_event():
    """
    Returns (subject, date_str, attendees, error).
    error is a machine-readable string: 'NOT_FOUND', 'AUTH', 'NETWORK', or the message.
    """
    cfg = _load_config()
    keyword = cfg.get("event_keyword", "DOXEE")
    cached_id = cfg.get("event_id")

    try:
        ev = gc.get_event(keyword, cached_id)
    except gc.NotFoundError:
        return None, None, [], "NOT_FOUND"
    except gc.AuthError as e:
        return None, None, [], f"AUTH: {e}"
    except Exception as e:
        return None, None, [], f"{type(e).__name__}: {e}"

    return ev["subject"], _fmt_iso(ev["start_iso"]), ev["attendees"], None


def add_attendee(name, email):
    cfg = _load_config()
    event_id = cfg.get("event_id")
    if not event_id:
        ev = gc.get_event(cfg.get("event_keyword", "DOXEE"), None)
        event_id = ev["id"]
    gc.add_attendee(event_id, name, email)


def load_invitati():
    if not INVITATI.exists():
        return []
    try:
        return json.loads(INVITATI.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


def save_invitati(data):
    INVITATI.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def _counts(attendees):
    c = {s: 0 for s in STATUS_COLORS}
    for a in attendees:
        if a["status"] in c:
            c[a["status"]] += 1
    return c


def build_xlsx(subject, date_str, attendees):
    """Builds an .xlsx file with the filtered attendees. Returns the bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "RSVP"

    # Header with event info
    ws["A1"] = subject or "Doxee Day"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    if date_str:
        ws["A2"] = f"Date: {date_str}"
        ws["A2"].font = Font(italic=True, color="6b7280")
        ws.merge_cells("A2:D2")

    ws["A3"] = f"Exported on {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(italic=True, color="9ca3af", size=10)
    ws.merge_cells("A3:D3")

    # Table header
    headers = ["Name", "Email", "Status", "Source"]
    header_row = 5
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1f2937")
        c.alignment = Alignment(horizontal="left")

    # Colors per status
    status_fill = {
        "Accepted":  "dcfce7",
        "Declined":  "fee2e2",
        "Tentative": "fef3c7",
        "Pending":   "f3f4f6",
    }

    for i, a in enumerate(attendees, start=header_row + 1):
        ws.cell(row=i, column=1, value=a.get("name", ""))
        ws.cell(row=i, column=2, value=a.get("email", ""))
        sc = ws.cell(row=i, column=3, value=a.get("status", ""))
        ws.cell(row=i, column=4, value=a.get("source", ""))
        if a.get("status") in status_fill:
            sc.fill = PatternFill("solid", fgColor=status_fill[a["status"]])

    # Column widths
    widths = {"A": 28, "B": 36, "C": 14, "D": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


_cache = {"data": None, "ts": 0}
_cache_lock = threading.Lock()
CACHE_TTL = 7


def get_data(force=False):
    with _cache_lock:
        if not force and _cache["data"] and (time.time() - _cache["ts"]) < CACHE_TTL:
            return _cache["data"]

    subject, date_str, attendees, err = read_event()

    if err == "NOT_FOUND":
        result = {
            "ok": False,
            "error": (
                f'No event with "{_event_kw()}" in the title found on Exchange.\n'
                f'Check that the event exists in your calendar, or change '
                f'"event_keyword" in config.json to match the exact title.'
            ),
        }
    elif err and err.startswith("AUTH:"):
        result = {
            "ok": False,
            "error": (
                f"Microsoft authentication failed: {err[5:].strip()}\n"
                f"Delete .token_cache.json and restart the server to log in again."
            ),
        }
    elif err:
        result = {
            "ok": False,
            "error": f"Errore Graph: {err}",
        }
    else:
        result = {
            "ok":        True,
            "subject":   subject or _event_kw(),
            "date":      date_str or "",
            "attendees": attendees,
            "counts":    _counts(attendees),
            "generated": datetime.now().strftime("%d/%m/%Y %H:%M"),
        }

    with _cache_lock:
        _cache["data"] = result
        _cache["ts"] = time.time()
    return result


class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        pass

    def _json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _html(self, content):
        body = content.encode()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _body(self):
        n = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(n)) if n else {}

    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/":
            self._html(HTML)
        elif path == "/api/data":
            self._json(get_data())
        elif path == "/api/debug/inbox":
            import requests as _req
            try:
                h = gc._h()
                r1 = _req.get(f"{gc.GRAPH}/me/messages", headers=h, params={
                    "$filter":  "isof('microsoft.graph.eventMessageResponse')",
                    "$top":     "10",
                    "$orderby": "receivedDateTime desc",
                }, timeout=20)
                r2 = _req.get(f"{gc.GRAPH}/me/messages", headers=h, params={
                    "$select":  "subject,from,receivedDateTime",
                    "$top":     "20",
                    "$orderby": "receivedDateTime desc",
                }, timeout=20)
                self._json({
                    "eventMessageResponse_only": r1.json().get("value", []) if r1.ok else {"error": r1.text},
                    "all_recent": r2.json().get("value", []) if r2.ok else {"error": r2.text},
                })
            except Exception as e:
                self._json({"error": str(e)}, 500)

        elif path == "/api/export":
            try:
                qs = parse_qs(urlparse(self.path).query)
                statuses = [s for s in qs.get("status", [""])[0].split(",") if s]
                data = get_data()
                if not data.get("ok"):
                    self._json({"ok": False, "error": data.get("error")}, 500)
                    return
                attendees = data["attendees"]
                if statuses:
                    attendees = [a for a in attendees if a.get("status") in statuses]
                xlsx = build_xlsx(data.get("subject", ""), data.get("date", ""), attendees)
                stamp = datetime.now().strftime("%Y%m%d_%H%M")
                tag = "_".join(statuses).lower() if statuses else "all"
                fname = f"rsvp_{tag}_{stamp}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type",
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                self.send_header("Content-Length", str(len(xlsx)))
                self.end_headers()
                self.wfile.write(xlsx)
            except Exception as e:
                self._json({"ok": False, "error": f"{type(e).__name__}: {e}"}, 500)
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        path = urlparse(self.path).path

        if path == "/api/invite":
            try:
                b     = self._body()
                name  = b.get("name", "").strip()
                email = b.get("email", "").strip().lower()

                if not name or not email or "@" not in email:
                    self._json({"ok": False, "error": "Nome e email valida richiesti."}, 400)
                    return

                add_attendee(name, email)

                with _cache_lock:
                    _cache["ts"] = 0

                self._json({"ok": True, "message": f"Invitation sent to {name}"})

            except gc.DuplicateError:
                self._json({"ok": False, "error": f"{email} is already invited to the event."}, 400)
            except gc.NotFoundError:
                self._json({"ok": False, "error": f'No event "{_event_kw()}" on Exchange.'}, 400)
            except gc.AuthError as e:
                self._json({"ok": False, "error": f"Auth failed: {e}"}, 401)
            except Exception as e:
                self._json({"ok": False, "error": f"{type(e).__name__}: {e}"}, 500)

        elif path == "/api/status":
            try:
                b      = self._body()
                email  = b.get("email", "").strip().lower()
                status = b.get("status", "").strip()
                name   = b.get("name", email).strip()

                if status not in STATUS_COLORS:
                    self._json({"ok": False, "error": "Invalid status."}, 400)
                    return

                invitati = load_invitati()
                found = False
                for inv in invitati:
                    if inv["email"].strip().lower() == email:
                        inv["status"] = status
                        found = True
                        break
                if not found:
                    invitati.append({
                        "name":       name,
                        "email":      email,
                        "status":     status,
                        "invited_at": datetime.now().isoformat(),
                    })
                save_invitati(invitati)
                with _cache_lock:
                    _cache["ts"] = 0
                self._json({"ok": True})
            except Exception as e:
                self._json({"ok": False, "error": str(e)}, 500)

        elif path == "/api/refresh":
            with _cache_lock:
                _cache["ts"] = 0
            self._json(get_data(force=True))

        elif path == "/api/reset":
            save_invitati([])
            with _cache_lock:
                _cache["data"] = None
                _cache["ts"] = 0
            self._json(get_data(force=True))

        else:
            self.send_response(404)
            self.end_headers()


HTML = """<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Event Mailer — RSVP</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,system-ui,"Segoe UI",sans-serif;background:#f3f4f6;color:#111827;min-height:100vh}

header{background:#1f2937;color:white;padding:18px 32px;display:flex;align-items:center;justify-content:space-between;gap:16px}
header h1{font-size:19px;font-weight:600}
header .sub{font-size:12px;opacity:.5;margin-top:3px}

.btn{border:none;padding:8px 16px;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;transition:background .15s}
.btn-green{background:#22c55e;color:white}.btn-green:hover{background:#16a34a}
.btn-ghost{background:transparent;border:1px solid rgba(255,255,255,.3);color:white}.btn-ghost:hover{background:rgba(255,255,255,.1)}
.btn:disabled{opacity:.4;cursor:not-allowed}

main{max-width:980px;margin:0 auto;padding:28px 32px}

.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:28px}
.card{color:white;padding:20px;border-radius:10px;text-align:center;cursor:pointer;
  transition:transform .12s, box-shadow .12s, opacity .15s;
  position:relative;user-select:none;outline:3px solid transparent;outline-offset:2px}
.card:hover{transform:translateY(-2px);box-shadow:0 6px 16px rgba(0,0,0,.12)}
.card.dim{opacity:.45}
.card.active{outline-color:#111827;box-shadow:0 6px 16px rgba(0,0,0,.18)}
.card .num{font-size:42px;font-weight:700;line-height:1}
.card .lbl{font-size:11px;text-transform:uppercase;letter-spacing:.07em;margin-top:6px;opacity:.9}
.card .check{position:absolute;top:8px;right:10px;font-size:14px;opacity:0;transition:opacity .15s}
.card.active .check{opacity:1}

.btn-blue{background:#2563eb;color:white}.btn-blue:hover{background:#1d4ed8}
.btn-gray{background:#e5e7eb;color:#374151}.btn-gray:hover{background:#d1d5db}

.toolbar{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.filter-info{font-size:12px;color:#6b7280}

.panel{background:white;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:22px}
.ph{padding:14px 20px;border-bottom:1px solid #e5e7eb;display:flex;align-items:center;justify-content:space-between}
.ph h2{font-size:14px;font-weight:600}
.pb{padding:20px}

.form-row{display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap}
.field{display:flex;flex-direction:column;gap:4px;flex:1;min-width:160px}
.field label{font-size:11px;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.04em}
.field input{padding:8px 12px;border:1px solid #d1d5db;border-radius:6px;font-size:14px;outline:none;transition:border .15s}
.field input:focus{border-color:#22c55e;box-shadow:0 0 0 3px rgba(34,197,94,.15)}

.toast{margin-top:12px;padding:10px 14px;border-radius:6px;font-size:13px;display:none}
.tok-ok{background:#dcfce7;color:#166534}.tok-err{background:#fee2e2;color:#991b1b}

.error-box{background:#fee2e2;border:1px solid #fca5a5;border-radius:8px;padding:16px 20px;margin-bottom:22px;font-size:13px;color:#991b1b;white-space:pre-line;display:none}

table{width:100%;border-collapse:collapse}
th,td{padding:11px 16px;text-align:left;border-bottom:1px solid #e5e7eb;font-size:13px}
th{font-size:11px;text-transform:uppercase;letter-spacing:.05em;color:#6b7280;font-weight:600;background:#f9fafb}
tbody tr:hover{background:#f9fafb}
tbody tr:last-child td{border-bottom:none}

.badge{display:inline-block;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700;color:white}
.badge.clickable{cursor:pointer;}.badge.clickable:hover{opacity:.8}
.src{font-size:10px;color:#9ca3af;margin-left:6px}

.menu{position:fixed;background:white;border:1px solid #e5e7eb;border-radius:8px;
  box-shadow:0 4px 16px rgba(0,0,0,.12);z-index:999;overflow:hidden;min-width:150px;display:none}
.mi{padding:9px 14px;font-size:13px;cursor:pointer;display:flex;align-items:center;gap:8px}
.mi:hover{background:#f3f4f6}
.md{width:8px;height:8px;border-radius:50%;flex-shrink:0}

.live-dot{display:inline-block;width:7px;height:7px;border-radius:50%;background:#22c55e;margin-right:6px;animation:pulse 2s infinite;vertical-align:middle}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}

.empty{text-align:center;padding:44px;color:#9ca3af;font-size:14px}
footer{text-align:center;color:#9ca3af;font-size:12px;margin-top:20px;padding-bottom:32px}
</style>
</head>
<body>

<header>
  <div>
    <h1 id="ev-title">Event Mailer — RSVP</h1>
    <div class="sub" id="ev-date"></div>
  </div>
  <div style="display:flex;gap:8px">
    <button class="btn btn-ghost" onclick="refresh()">↻ Refresh</button>
    <button class="btn btn-ghost" style="color:#ef4444;border-color:#fca5a5" onclick="resetAll()">⊘ Reset</button>
  </div>
</header>

<main>
  <div class="error-box" id="err-box"></div>

  <div class="cards">
    <div class="card" data-status="Accepted" style="background:#22c55e" onclick="toggleFilter('Accepted')">
      <span class="check">✓</span><div class="num" id="c-acc">—</div><div class="lbl">Accepted</div></div>
    <div class="card" data-status="Declined" style="background:#ef4444" onclick="toggleFilter('Declined')">
      <span class="check">✓</span><div class="num" id="c-dec">—</div><div class="lbl">Declined</div></div>
    <div class="card" data-status="Tentative" style="background:#eab308" onclick="toggleFilter('Tentative')">
      <span class="check">✓</span><div class="num" id="c-ten">—</div><div class="lbl">Tentative</div></div>
    <div class="card" data-status="Pending" style="background:#9ca3af" onclick="toggleFilter('Pending')">
      <span class="check">✓</span><div class="num" id="c-pen">—</div><div class="lbl">Pending</div></div>
  </div>

  <div class="panel">
    <div class="ph"><h2>Add guest</h2></div>
    <div class="pb">
      <div class="form-row">
        <div class="field"><label>Name</label><input id="f-name" type="text" placeholder="Mario Rossi"/></div>
        <div class="field"><label>Email</label><input id="f-email" type="email" placeholder="mario@company.com"/></div>
        <button class="btn btn-green" id="btn-inv" onclick="sendInvite()">Send invitation</button>
      </div>
      <div id="toast" class="toast"></div>
    </div>
  </div>

  <div class="panel">
    <div class="ph">
      <h2>Attendees <span id="total" style="color:#9ca3af;font-weight:400;font-size:12px;margin-left:6px"></span></h2>
      <div class="toolbar">
        <span class="filter-info" id="filter-info"></span>
        <button class="btn btn-gray" id="btn-clear" onclick="clearFilters()" style="display:none">Clear filters</button>
        <button class="btn btn-blue" onclick="exportXlsx()">⬇ Export XLSX</button>
        <span style="font-size:12px;color:#6b7280"><span class="live-dot"></span>live</span>
      </div>
    </div>
    <div id="tbl"></div>
  </div>

  <footer id="foot"></footer>
</main>

<div class="menu" id="menu"></div>

<script>
const SC = {Accepted:'#22c55e',Declined:'#ef4444',Tentative:'#eab308',Pending:'#9ca3af'};
let _menuEmail = '', _menuName = '';
let _lastData = null;
let _activeFilters = new Set();

function toggleFilter(status){
  if(_activeFilters.has(status)) _activeFilters.delete(status);
  else _activeFilters.add(status);
  updateFilterUI();
  if(_lastData) render(_lastData);
}

function clearFilters(){
  _activeFilters.clear();
  updateFilterUI();
  if(_lastData) render(_lastData);
}

function updateFilterUI(){
  const has = _activeFilters.size > 0;
  document.querySelectorAll('.card').forEach(c=>{
    const s = c.dataset.status;
    if(!s) return;
    c.classList.toggle('active', _activeFilters.has(s));
    c.classList.toggle('dim', has && !_activeFilters.has(s));
  });
  document.getElementById('btn-clear').style.display = has ? 'inline-block' : 'none';
  const info = document.getElementById('filter-info');
  info.textContent = has ? `Filtro: ${[..._activeFilters].join(', ')}` : '';
}

function exportXlsx(){
  const params = _activeFilters.size ? `?status=${[..._activeFilters].join(',')}` : '';
  window.location.href = '/api/export' + params;
}

function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}

async function loadData(){
  try{
    const d = await fetch('/api/data').then(r=>r.json());
    render(d);
  }catch(e){console.error(e)}
}

async function refresh(){
  try{
    const d = await fetch('/api/refresh',{method:'POST'}).then(r=>r.json());
    render(d);
  }catch(e){console.error(e)}
}

async function resetAll(){
  if(!confirm('Empty invitati.json and reload from Exchange?')) return;
  try{
    const d = await fetch('/api/reset',{method:'POST'}).then(r=>r.json());
    render(d);
  }catch(e){console.error(e)}
}

function clearDisplay(){
  ['c-acc','c-dec','c-ten','c-pen'].forEach(id=>document.getElementById(id).textContent='—');
  document.getElementById('total').textContent='';
  document.getElementById('tbl').innerHTML='';
  document.getElementById('foot').textContent='';
  document.getElementById('ev-title').textContent='Event Mailer — RSVP';
  document.getElementById('ev-date').textContent='';
}

function render(d){
  _lastData = d;
  const errBox = document.getElementById('err-box');
  if(!d.ok){
    clearDisplay();
    errBox.textContent = d.error || 'Errore sconosciuto.';
    errBox.style.display = 'block';
    return;
  }
  errBox.style.display = 'none';

  document.getElementById('ev-title').textContent = d.subject || 'Doxee Day';
  document.getElementById('ev-date').textContent  = d.date || '';
  document.getElementById('c-acc').textContent = d.counts.Accepted;
  document.getElementById('c-dec').textContent = d.counts.Declined;
  document.getElementById('c-ten').textContent = d.counts.Tentative;
  document.getElementById('c-pen').textContent = d.counts.Pending;

  const filtered = _activeFilters.size
    ? d.attendees.filter(a => _activeFilters.has(a.status))
    : d.attendees;

  document.getElementById('total').textContent =
    _activeFilters.size ? `(${filtered.length} of ${d.attendees.length})` : `(${d.attendees.length})`;
  document.getElementById('foot').textContent  = `Updated ${d.generated}`;

  if(!filtered.length){
    const msg = _activeFilters.size
      ? 'No attendees match the selected filters.'
      : 'No attendees yet. Add the first guest above.';
    document.getElementById('tbl').innerHTML = `<div class="empty">${msg}</div>`;
    return;
  }

  const order = ['Accepted','Tentative','Pending','Declined'];
  const sorted = [...filtered].sort((a,b)=>order.indexOf(a.status)-order.indexOf(b.status));

  document.getElementById('tbl').innerHTML = `<table>
    <thead><tr><th>Name</th><th>Email</th><th>Status</th></tr></thead>
    <tbody>${sorted.map(a=>{
      const manual = a.source === 'manual';
      const badge = manual
        ? `<span class="badge clickable" style="background:${SC[a.status]||'#6b7280'}"
             onclick="openMenu(event,'${esc(a.email)}','${esc(a.name)}')"
             title="Click to update">${esc(a.status)} ▾</span>`
        : `<span class="badge" style="background:${SC[a.status]||'#6b7280'}">${esc(a.status)}</span>`;
      return `<tr>
        <td>${esc(a.name)}${manual?'<span class="src">(manual)</span>':''}</td>
        <td style="color:#6b7280;font-size:12px">${esc(a.email)}</td>
        <td>${badge}</td>
      </tr>`;
    }).join('')}
    </tbody></table>`;
}

async function sendInvite(){
  const name  = document.getElementById('f-name').value.trim();
  const email = document.getElementById('f-email').value.trim();
  const btn   = document.getElementById('btn-inv');
  if(!name||!email){toast('Enter name and email.',false);return}
  btn.disabled=true; btn.textContent='Sending…';
  try{
    const r = await fetch('/api/invite',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({name,email})}).then(r=>r.json());
    if(r.ok){
      toast(r.message,true);
      document.getElementById('f-name').value='';
      document.getElementById('f-email').value='';
      setTimeout(loadData,1500);
    }else{
      console.error('[RSVP] /api/invite error:', r.error);
      toast(r.error,false,true);
    }
  }catch(e){
    console.error('[RSVP] fetch error:', e);
    toast('Connection error.',false,true);
  }
  btn.disabled=false; btn.textContent='Send invitation';
}

function closeToast(){document.getElementById('toast').style.display='none';}

function toast(msg,ok,sticky=false){
  const t=document.getElementById('toast');
  t.innerHTML=msg+(sticky?' <span onclick="closeToast()" style="float:right;cursor:pointer;margin-left:12px;font-weight:700">&#x2715;</span>':'');
  t.className='toast '+(ok?'tok-ok':'tok-err');
  t.style.display='block';
  clearTimeout(t._t);
  if(!sticky) t._t=setTimeout(()=>t.style.display='none',4000);
}

function openMenu(e, email, name){
  e.stopPropagation();
  _menuEmail = email; _menuName = name;
  const m = document.getElementById('menu');
  m.innerHTML = Object.entries(SC).map(([s,c])=>
    `<div class="mi" onclick="setStatus('${s}')">
      <span class="md" style="background:${c}"></span>${s}
    </div>`).join('');
  m.style.display = 'block';
  m.style.left = Math.min(e.clientX, window.innerWidth-160)+'px';
  m.style.top  = (e.clientY+6)+'px';
}

async function setStatus(status){
  document.getElementById('menu').style.display = 'none';
  await fetch('/api/status',{method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({email:_menuEmail, name:_menuName, status})});
  loadData();
}

document.addEventListener('click', ()=>document.getElementById('menu').style.display='none');
document.getElementById('f-email').addEventListener('keydown',e=>{if(e.key==='Enter')sendInvite()});

loadData();
setInterval(loadData, 8000);
</script>
</body>
</html>"""


def main():
    # First-run wizard (auto-detects .ics, collects credentials, stores secret in the keychain).
    if setup.needs_setup():
        setup.run()

    url = f"http://localhost:{PORT}"

    print("[event-mailer] Checking Microsoft Graph authentication...")
    try:
        gc._ensure_token()
        print("[event-mailer] Auth OK.")
    except gc.AuthError as e:
        print(f"[event-mailer] AUTH ERROR: {e}")
        print("[event-mailer] Delete .token_cache.json and try again.")
        return

    try:
        httpd = HTTPServer(("localhost", PORT), Handler)
    except OSError:
        print(f"[ERROR] Port {PORT} already in use. Close the other instance and try again.")
        return

    print(f"[event-mailer] Panel started -> {url}")
    print("[event-mailer] Ctrl+C to stop.")

    def _open():
        time.sleep(0.6)
        webbrowser.open(url)
    threading.Thread(target=_open, daemon=True).start()

    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n[event-mailer] Fermato.")
        httpd.shutdown()


if __name__ == "__main__":
    main()
